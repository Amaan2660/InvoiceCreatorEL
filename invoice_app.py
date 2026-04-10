# Updated Streamlit Invoice App with Single + Bulk Invoice Creation,
# Gmail sending after generation, persistent downloads, and Nordea/Revolut bank selection

import datetime
import pandas as pd
import base64
import zipfile
import smtplib
from fpdf import FPDF
from io import BytesIO
from email.message import EmailMessage

import streamlit as st
from sqlalchemy import create_engine, Column, String, Integer, Boolean
from sqlalchemy.orm import declarative_base, sessionmaker
from openpyxl import load_workbook
from openpyxl.styles import Font

# ------------------- PAGE / SESSION SETUP -------------------
st.set_page_config(page_title="InvoiceCreatorEL", layout="centered")

SESSION_DEFAULTS = {
    "single_generated_pdf_bytes": None,
    "single_generated_pdf_name": None,
    "single_generated_spec_bytes": None,
    "single_generated_spec_name": None,
    "single_generated_preview_df": None,
    "bulk_preview_df": None,
    "bulk_results": [],
    "bulk_zip_bytes": None,
    "bulk_zip_name": None,
    "bulk_last_run_summary": None,
}

for key, value in SESSION_DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = value

# ------------------- DATABASE SETUP -------------------
DB_URL = st.secrets["SUPABASE_DB_URL"]
engine = create_engine(DB_URL)
Base = declarative_base()
SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)

class Customer(Base):
    __tablename__ = "customers"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    address = Column(String)
    email = Column(String)
    contact = Column(String)
    vat = Column(String)
    is_company = Column(Boolean, default=True)

Base.metadata.create_all(bind=engine)

def add_customer(**kwargs):
    with SessionLocal() as session:
        session.add(Customer(**kwargs))
        session.commit()

def get_customers():
    with SessionLocal() as session:
        return session.query(Customer).order_by(Customer.name).all()

def update_customer(id, updates):
    with SessionLocal() as session:
        cust = session.get(Customer, id)
        if cust:
            for k, v in updates.items():
                setattr(cust, k, v)
            session.commit()

def delete_customer(id):
    with SessionLocal() as session:
        session.query(Customer).filter(Customer.id == id).delete()
        session.commit()

# ------------------- HELPERS -------------------
def convert_currency(amount_dkk, target_currency):
    rates = {
        "EUR": 7.5,
        "USD": 6.5,
        "GBP": 8.8
    }
    rate = rates.get(target_currency)
    return round(amount_dkk / rate, 2) if rate else amount_dkk

def get_currency_note(currency):
    rates = {
        "EUR": 7.5,
        "USD": 6.5,
        "GBP": 8.8
    }
    return f"{currency} (1 {currency} = {rates[currency]} DKK)" if currency in rates else currency

def get_bank_details(bank_choice):
    if bank_choice == "Nordea":
        return {
            "bank_name": "Nordea",
            "iban": "DK41 2000 9046 3317 85",
            "swift": "NDEADKKK",
            "reg_no": "2355",
            "account_no": "9046331785"
        }
    return {
        "bank_name": "Revolut",
        "iban": "LT87 3250 0345 4552 5735",
        "swift": "REVOLT21",
        "reg_no": "",
        "account_no": ""
    }

def customer_to_dict(customer):
    if customer is None:
        return {
            "id": None,
            "name": "",
            "address": "",
            "email": "",
            "contact": "",
            "vat": "",
            "is_company": True,
        }
    return {
        "id": customer.id,
        "name": customer.name or "",
        "address": customer.address or "",
        "email": customer.email or "",
        "contact": customer.contact or "",
        "vat": customer.vat or "",
        "is_company": bool(customer.is_company),
    }

def preview_pdf(bytes_pdf):
    b64 = base64.b64encode(bytes_pdf).decode()
    return f"<iframe src='data:application/pdf;base64,{b64}' width='700' height='900' type='application/pdf'></iframe>"

def preview_excel(df):
    return st.dataframe(df, use_container_width=True)

def normalize_name(value):
    return str(value).strip().lower() if pd.notna(value) else ""

def find_best_customer_match(customer_name, customers):
    target = normalize_name(customer_name)
    for customer in customers:
        if normalize_name(customer.name) == target:
            return customer
    for customer in customers:
        if target and (target in normalize_name(customer.name) or normalize_name(customer.name) in target):
            return customer
    return None

def read_excel_file(uploaded_file):
    if uploaded_file.name.endswith(".xls"):
        df = pd.read_excel(uploaded_file, header=1, engine="xlrd")
    else:
        df = pd.read_excel(uploaded_file, header=1, engine="openpyxl")
    return df

def clean_trip_dataframe(df):
    target_cols = ['Trip Date', 'Passenger', 'From', 'To', 'Customer', 'Cust. Ref.', 'Base Rate']
    missing_cols = [col for col in target_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns in Excel file: {', '.join(missing_cols)}")

    cleaned_df = df[target_cols].copy()
    cleaned_df = cleaned_df.dropna(subset=['Base Rate'])
    cleaned_df['Base Rate'] = cleaned_df['Base Rate'].astype(str).str.replace(',', '', regex=False)
    cleaned_df['Base Rate'] = pd.to_numeric(cleaned_df['Base Rate'], errors='coerce')
    cleaned_df = cleaned_df.dropna(subset=['Base Rate'])

    if not cleaned_df.empty and len(cleaned_df) > 1:
        last_value = cleaned_df['Base Rate'].iloc[-1]
        sum_except_last = cleaned_df['Base Rate'].iloc[:-1].sum()
        if abs(last_value - sum_except_last) < 1.0:
            cleaned_df = cleaned_df.iloc[:-1]

    cleaned_df = cleaned_df.reset_index(drop=True)
    return cleaned_df

def build_specification_workbook_bytes(cleaned_df):
    buffer = BytesIO()
    cleaned_df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold_font

    for col in ws.columns:
        values = [len(str(cell.value)) for cell in col if cell.value is not None]
        max_length = max(values) if values else 10
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    ws.append(["", "", "", "", "", "Total", cleaned_df['Base Rate'].sum()])

    final_buffer = BytesIO()
    wb.save(final_buffer)
    return final_buffer.getvalue()

def build_email_body(customer_name, invoice_number, due_date, currency):
    due_date_str = due_date.strftime("%d/%m/%Y") if hasattr(due_date, "strftime") else str(due_date)
    return (
        f"Dear {customer_name},\n\n"
        f"Please find attached invoice {invoice_number}.\n"
        f"Due date: {due_date_str}\n"
        f"Currency: {currency}\n\n"
        f"Best regards,\n"
        f"Limousine Service Xpress ApS"
    )

def send_email_gmail(to_email, subject, body, attachments):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = st.secrets["SMTP_SENDER"]
    msg["To"] = to_email
    msg.set_content(body)

    for attachment in attachments:
        msg.add_attachment(
            attachment["content"],
            maintype=attachment["maintype"],
            subtype=attachment["subtype"],
            filename=attachment["filename"]
        )

    with smtplib.SMTP(st.secrets["SMTP_HOST"], int(st.secrets["SMTP_PORT"])) as server:
        server.starttls()
        server.login(st.secrets["SMTP_USERNAME"], st.secrets["SMTP_PASSWORD"])
        server.send_message(msg)

# ------------------- PDF GENERATION -------------------
def generate_invoice_pdf(receiver, invoice_number, currency, description, total_amount, booking_count, due_date, bank_choice):
    pdf = FPDF()
    pdf.add_page()
    bank_details = get_bank_details(bank_choice)

    try:
        pdf.image("logo.png", x=10, y=8, w=40)
    except Exception:
        pass

    pdf.set_font("Helvetica", "B", 20)
    pdf.set_xy(150, 10)
    pdf.cell(0, 10, f"INVOICE {invoice_number}", ln=True)

    pdf.set_font("Helvetica", size=11)
    pdf.set_xy(10, 30)
    pdf.set_font("Helvetica", "B", 11)
    pdf.multi_cell(90, 6, "From:")
    pdf.set_font("Helvetica", "", 11)

    sender_text = f"""Limousine Service Xpress ApS
Industriholmen 82
2650 Hvidovre
Denmark
CVR/VAT: DK45247961
IBAN: {bank_details['iban']}
SWIFT: {bank_details['swift']}"""

    if bank_details["reg_no"]:
        sender_text += f"""
Reg Nr: {bank_details['reg_no']}
Konto Nr: {bank_details['account_no']}"""

    sender_text += """
Email: limoexpresscph@gmail.com"""

    pdf.multi_cell(90, 6, sender_text)

    pdf.set_xy(120, 30)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, "To:", ln=True)
    pdf.set_font("Helvetica", "", 11)

    pdf.set_x(120)
    if receiver.get("name"):
        pdf.cell(0, 6, receiver["name"], ln=True)
    if receiver.get("contact"):
        pdf.set_x(120)
        pdf.multi_cell(80, 6, f"Att: {receiver['contact']}")
    if receiver.get("address"):
        pdf.set_x(120)
        pdf.multi_cell(80, 6, receiver["address"])
    if receiver.get("vat") and receiver.get("is_company"):
        pdf.set_x(120)
        pdf.cell(0, 6, f"VAT No: {receiver['vat']}", ln=True)
    if receiver.get("email"):
        pdf.set_x(120)
        pdf.multi_cell(80, 6, f"Email: {receiver['email']}")

    pdf.set_xy(10, 100)
    today = datetime.date.today().strftime("%d/%m/%Y")
    due_date_fmt = due_date.strftime("%d/%m/%Y") if hasattr(due_date, "strftime") else str(due_date)
    currency_note = get_currency_note(currency)
    pdf.cell(0, 6, f"Invoice Date: {today}", ln=True)
    pdf.cell(0, 6, f"Due Date: {due_date_fmt}", ln=True)
    pdf.cell(0, 6, f"Currency: {currency_note}", ln=True)
    pdf.ln(10)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(120, 8, "Description", border=1)
    pdf.cell(30, 8, "Qty", border=1)
    pdf.cell(40, 8, "Total", border=1, ln=True)

    pdf.set_font("Helvetica", "", 11)
    pdf.cell(120, 8, description or "Transfers", border=1)
    pdf.cell(30, 8, str(booking_count), border=1)
    pdf.cell(40, 8, f"{total_amount:,.2f} {currency}", border=1, ln=True)

    pdf.ln(8)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(150, 8, "Subtotal:", border=0)
    pdf.cell(40, 8, f"{float(total_amount):,.2f} {currency}", ln=True)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(150, 8, "Total Amount Due:", border=0)
    pdf.cell(40, 8, f"{total_amount:,.2f} {currency}", ln=True)

    pdf.ln(10)
    pdf.set_font("Helvetica", "I", 10)
    pdf.cell(0, 6, f"Please add invoice number {invoice_number} as reference when making payment.", ln=True)

    if bank_choice == "Nordea":
        previous_auto_page_break = pdf.auto_page_break
        previous_bottom_margin = pdf.b_margin
        pdf.set_auto_page_break(False)
        pdf.set_y(-50)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(190, 8, "IMPORTANT: PLEASE USE OUR NEW BANK DETAILS", border=1, ln=True, align="C")
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(
            190,
            6,
            "Please make payment using the Nordea banking information stated on this invoice.",
            border=1,
            align="C"
        )
        pdf.set_auto_page_break(previous_auto_page_break, margin=previous_bottom_margin)

    return pdf.output(dest="S").encode("latin-1")

# ------------------- BULK HELPERS -------------------
def build_bulk_groups(cleaned_df, customers):
    groups = []
    for customer_name, group_df in cleaned_df.groupby("Customer", dropna=False):
        display_name = str(customer_name).strip() if pd.notna(customer_name) and str(customer_name).strip() else "Unknown Customer"
        matched_customer = find_best_customer_match(display_name, customers)

        groups.append({
            "group_customer_name": display_name,
            "matched_customer": matched_customer,
            "dataframe": group_df.reset_index(drop=True),
            "trip_count": int(group_df.shape[0]),
            "total_dkk": float(group_df["Base Rate"].sum()),
        })

    groups.sort(key=lambda x: x["group_customer_name"].lower())
    return groups

def create_zip_from_bulk_results(results):
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for result in results:
            zip_file.writestr(result["pdf_filename"], result["pdf_bytes"])
            if result["spec_bytes"] is not None:
                zip_file.writestr(result["spec_filename"], result["spec_bytes"])
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def validate_bulk_rows(rows):
    errors = []
    invoice_numbers = []

    for idx, row in enumerate(rows, start=1):
        if not row.get("include", True):
            continue

        invoice_number = str(row.get("invoice_number", "")).strip()
        if not invoice_number:
            errors.append(f"Row {idx}: Invoice number is required.")
        else:
            invoice_numbers.append(invoice_number)

        recipient_name = str(row.get("recipient_name", "")).strip()
        if not recipient_name:
            errors.append(f"Row {idx}: Recipient name is required.")

    duplicates = pd.Series(invoice_numbers).duplicated(keep=False) if invoice_numbers else pd.Series(dtype=bool)
    if len(invoice_numbers) > 0 and duplicates.any():
        duplicate_values = sorted(set([num for num, dup in zip(invoice_numbers, duplicates) if dup]))
        errors.append(f"Duplicate invoice numbers found: {', '.join(duplicate_values)}")

    return errors

def generate_single_invoice_package(
    receiver_dict,
    invoice_number,
    currency,
    bank_choice,
    invoice_purpose,
    due_date,
    cleaned_df
):
    booking_count = int(cleaned_df.shape[0])
    total_amount_dkk = float(cleaned_df["Base Rate"].sum())
    final_total = convert_currency(total_amount_dkk, currency) if currency != "DKK" else total_amount_dkk

    spec_bytes = build_specification_workbook_bytes(cleaned_df)
    spec_filename = f"SERVICE SPECIFICATION FOR INVOICE {invoice_number}.xlsx"

    pdf_bytes = generate_invoice_pdf(
        receiver=receiver_dict,
        invoice_number=invoice_number,
        currency=currency,
        description=invoice_purpose,
        total_amount=final_total,
        booking_count=booking_count,
        due_date=due_date,
        bank_choice=bank_choice
    )
    pdf_filename = f"Invoice {invoice_number} for {receiver_dict['name']}.pdf"

    return {
        "pdf_bytes": pdf_bytes,
        "pdf_filename": pdf_filename,
        "spec_bytes": spec_bytes,
        "spec_filename": spec_filename,
        "booking_count": booking_count,
        "total_amount_dkk": total_amount_dkk,
        "final_total": final_total,
        "preview_df": cleaned_df.copy(),
    }

# ------------------- UI -------------------
st.title("📄 Invoice Creator EL")
tab1, tab2 = st.tabs(["🧾 Create Invoice", "👥 Manage Customers"])

with tab1:
    st.subheader("Create Invoice")
    customers = get_customers()

    creation_mode = st.radio(
        "Invoice Mode",
        ["Single Invoice", "Bulk Invoices"],
        horizontal=True
    )

    if creation_mode == "Single Invoice":
        receiver = st.selectbox("Select Customer", customers, format_func=lambda x: x.name if x else "")
        invoice_number = st.text_input("Invoice Number")
        currency = st.selectbox("Currency", ["DKK", "EUR", "USD", "GBP"], key="single_currency")
        bank_choice = st.selectbox("Bank for payment", ["Nordea", "Revolut"], index=0, key="single_bank")
        invoice_purpose = st.text_input("Invoice Description (e.g. Transfers in May 2025)")
        due_date = st.date_input("Due Date", key="single_due_date")
        mode = st.radio("Select Amount Mode", ["Manual", "Auto from Excel"], key="single_amount_mode")

        uploaded = st.file_uploader("Upload Excel File", type=["xls", "xlsx"], key="single_file")
        total_amount_dkk = 0.0
        booking_count = 0
        cleaned_df = pd.DataFrame()

        if uploaded:
            try:
                df = read_excel_file(uploaded)
                cleaned_df = clean_trip_dataframe(df)

                if mode == "Auto from Excel":
                    booking_count = int(cleaned_df.shape[0])
                    total_amount_dkk = float(cleaned_df['Base Rate'].sum())
            except Exception as e:
                st.error(f"Could not read Excel file: {e}")

        if mode == "Manual":
            total_amount_dkk = st.number_input("Manual Total Amount", min_value=0.0, step=100.0)
            booking_count = st.number_input("Manual Number of Bookings", min_value=0)

        if st.button("Generate Invoice", key="generate_single_invoice"):
            if not receiver or not invoice_number:
                st.error("Customer and Invoice Number are required.")
            else:
                spec_bytes = None
                spec_name = None
                preview_df = None

                if not cleaned_df.empty:
                    spec_bytes = build_specification_workbook_bytes(cleaned_df)
                    spec_name = f"SERVICE SPECIFICATION FOR INVOICE {invoice_number}.xlsx"
                    preview_df = cleaned_df.copy()

                receiver_dict = customer_to_dict(receiver)
                final_total = convert_currency(total_amount_dkk, currency) if mode == "Auto from Excel" and currency != "DKK" else total_amount_dkk

                pdf_bytes = generate_invoice_pdf(
                    receiver=receiver_dict,
                    invoice_number=invoice_number,
                    currency=currency,
                    description=invoice_purpose,
                    total_amount=final_total,
                    booking_count=booking_count,
                    due_date=due_date,
                    bank_choice=bank_choice
                )

                st.session_state.single_generated_pdf_bytes = pdf_bytes
                st.session_state.single_generated_pdf_name = f"Invoice {invoice_number} for {receiver.name}.pdf"
                st.session_state.single_generated_spec_bytes = spec_bytes
                st.session_state.single_generated_spec_name = spec_name
                st.session_state.single_generated_preview_df = preview_df

        if st.session_state.single_generated_preview_df is not None:
            preview_excel(st.session_state.single_generated_preview_df)

        if st.session_state.single_generated_spec_bytes is not None:
            st.download_button(
                "⬇️ Download Specification XLSX",
                data=st.session_state.single_generated_spec_bytes,
                file_name=st.session_state.single_generated_spec_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_single_spec"
            )

        if st.session_state.single_generated_pdf_bytes is not None:
            st.markdown(preview_pdf(st.session_state.single_generated_pdf_bytes), unsafe_allow_html=True)
            st.download_button(
                "⬇️ Download PDF Invoice",
                data=st.session_state.single_generated_pdf_bytes,
                file_name=st.session_state.single_generated_pdf_name,
                mime="application/pdf",
                key="download_single_pdf"
            )

    else:
        st.markdown("### Bulk Invoice Creation")
        st.caption("Upload one Excel file with many trips. The app will group rows by Customer.")
        st.info("Invoices are generated first. After generation, you can choose which ones to email and send them together.")

        bulk_uploaded = st.file_uploader("Upload Bulk Excel File", type=["xls", "xlsx"], key="bulk_file")
        default_due_date = st.date_input("Default Due Date", value=datetime.date.today(), key="bulk_default_due")
        default_currency = st.selectbox("Default Currency", ["DKK", "EUR", "USD", "GBP"], key="bulk_default_currency")
        default_bank = st.selectbox("Default Bank", ["Nordea", "Revolut"], index=0, key="bulk_default_bank")
        default_description = st.text_input(
            "Default Invoice Description",
            value="Transfers",
            key="bulk_default_description"
        )
        starting_invoice_number = st.text_input(
            "Starting Invoice Number (optional, used to prefill)",
            key="bulk_start_invoice_number"
        )

        if bulk_uploaded:
            try:
                bulk_df_raw = read_excel_file(bulk_uploaded)
                bulk_cleaned_df = clean_trip_dataframe(bulk_df_raw)
                st.session_state.bulk_preview_df = bulk_cleaned_df.copy()

                bulk_groups = build_bulk_groups(bulk_cleaned_df, customers)

                if not bulk_groups:
                    st.warning("No customer groups found in the file.")
                else:
                    st.markdown("### Customer Groups")

                    if starting_invoice_number.strip().isdigit():
                        start_num = int(starting_invoice_number.strip())
                    else:
                        start_num = None

                    bulk_rows = []
                    preview_rows = []

                    for idx, group in enumerate(bulk_groups):
                        group_name = group["group_customer_name"]
                        matched_customer = group["matched_customer"]
                        default_customer = matched_customer
                        default_invoice = str(start_num + idx) if start_num is not None else ""
                        default_email = matched_customer.email if matched_customer and matched_customer.email else ""
                        default_name = matched_customer.name if matched_customer else group_name
                        default_address = matched_customer.address if matched_customer else ""
                        default_contact = matched_customer.contact if matched_customer else ""
                        default_vat = matched_customer.vat if matched_customer else ""
                        default_is_company = bool(matched_customer.is_company) if matched_customer else True

                        with st.expander(f"{group_name} — {group['trip_count']} trips — {group['total_dkk']:,.2f} DKK", expanded=False):
                            include = st.checkbox("Include this invoice", value=True, key=f"bulk_include_{idx}")
                            send_email_flag = st.checkbox("Mark for email sending", value=bool(default_email), key=f"bulk_send_email_{idx}")

                            match_key = f"bulk_match_customer_{idx}"
                            name_key = f"bulk_recipient_name_{idx}"
                            email_key = f"bulk_email_{idx}"
                            address_key = f"bulk_address_{idx}"
                            contact_key = f"bulk_contact_{idx}"
                            vat_key = f"bulk_vat_{idx}"
                            company_key = f"bulk_is_company_{idx}"

                            options_list = [None] + customers
                            default_index = options_list.index(default_customer) if default_customer in customers else 0

                            chosen_db_customer = st.selectbox(
                                "Match saved customer",
                                options=options_list,
                                index=default_index,
                                format_func=lambda x: x.name if x else "No match / custom values",
                                key=match_key
                            )

                            if f"{match_key}_last_id" not in st.session_state:
                                st.session_state[f"{match_key}_last_id"] = chosen_db_customer.id if chosen_db_customer else None

                            current_id = chosen_db_customer.id if chosen_db_customer else None
                            previous_id = st.session_state[f"{match_key}_last_id"]

                            if previous_id != current_id:
                                if chosen_db_customer:
                                    st.session_state[name_key] = chosen_db_customer.name or ""
                                    st.session_state[email_key] = chosen_db_customer.email or ""
                                    st.session_state[address_key] = chosen_db_customer.address or ""
                                    st.session_state[contact_key] = chosen_db_customer.contact or ""
                                    st.session_state[vat_key] = chosen_db_customer.vat or ""
                                    st.session_state[company_key] = bool(chosen_db_customer.is_company)
                                else:
                                    st.session_state[name_key] = group_name
                                    st.session_state[email_key] = ""
                                    st.session_state[address_key] = ""
                                    st.session_state[contact_key] = ""
                                    st.session_state[vat_key] = ""
                                    st.session_state[company_key] = True

                                st.session_state[f"{match_key}_last_id"] = current_id

                            if name_key not in st.session_state:
                                st.session_state[name_key] = default_name
                            if email_key not in st.session_state:
                                st.session_state[email_key] = default_email
                            if address_key not in st.session_state:
                                st.session_state[address_key] = default_address
                            if contact_key not in st.session_state:
                                st.session_state[contact_key] = default_contact
                            if vat_key not in st.session_state:
                                st.session_state[vat_key] = default_vat
                            if company_key not in st.session_state:
                                st.session_state[company_key] = default_is_company

                            recipient_name = st.text_input("Recipient Name", key=name_key)
                            recipient_email = st.text_input("Recipient Email", key=email_key)
                            recipient_address = st.text_input("Recipient Address", key=address_key)
                            recipient_contact = st.text_input("Recipient Contact", key=contact_key)
                            recipient_vat = st.text_input("Recipient VAT Number", key=vat_key)
                            recipient_is_company = st.checkbox("Recipient Is Company", key=company_key)

                            invoice_key = f"bulk_invoice_number_{idx}"
                            start_seed_key = f"bulk_invoice_seed_{idx}"

                            if invoice_key not in st.session_state:
                                st.session_state[invoice_key] = default_invoice

                            if start_seed_key not in st.session_state:
                                st.session_state[start_seed_key] = starting_invoice_number.strip()

                            current_seed = starting_invoice_number.strip()
                            previous_seed = st.session_state[start_seed_key]

                            if current_seed != previous_seed:
                                if current_seed.isdigit() and not str(st.session_state[invoice_key]).strip():
                                    st.session_state[invoice_key] = str(int(current_seed) + idx)
                                st.session_state[start_seed_key] = current_seed

                            desc_key = f"bulk_description_{idx}"
                            desc_seed_key = f"bulk_description_seed_{idx}"

                            if desc_key not in st.session_state:
                                st.session_state[desc_key] = default_description

                            if desc_seed_key not in st.session_state:
                                st.session_state[desc_seed_key] = default_description

                            if default_description != st.session_state[desc_seed_key]:
                                if not st.session_state[desc_key]:
                                    st.session_state[desc_key] = default_description
                                st.session_state[desc_seed_key] = default_description

                            col_a, col_b, col_c = st.columns(3)
                            with col_a:
                                invoice_number_val = st.text_input("Invoice Number", key=invoice_key)
                            with col_b:
                                currency_val = st.selectbox(
                                    "Currency",
                                    ["DKK", "EUR", "USD", "GBP"],
                                    index=["DKK", "EUR", "USD", "GBP"].index(default_currency),
                                    key=f"bulk_currency_{idx}"
                                )
                            with col_c:
                                bank_val = st.selectbox(
                                    "Bank",
                                    ["Nordea", "Revolut"],
                                    index=["Nordea", "Revolut"].index(default_bank),
                                    key=f"bulk_bank_{idx}"
                                )

                            due_date_val = st.date_input("Due Date", value=default_due_date, key=f"bulk_due_date_{idx}")
                            description_val = st.text_input("Description", key=desc_key)

                            preview_rows.append({
                                "Include": include,
                                "Customer in File": group_name,
                                "Matched DB Customer": chosen_db_customer.name if chosen_db_customer else "",
                                "Trips": group["trip_count"],
                                "Total DKK": round(group["total_dkk"], 2),
                                "Invoice Number": invoice_number_val,
                                "Currency": currency_val,
                                "Email": recipient_email,
                                "Mark for Sending": send_email_flag,
                                "Description": description_val,
                            })

                            bulk_rows.append({
                                "include": include,
                                "send_email": send_email_flag,
                                "group_name": group_name,
                                "chosen_db_customer": chosen_db_customer,
                                "recipient_name": recipient_name,
                                "email": recipient_email,
                                "address": recipient_address,
                                "contact": recipient_contact,
                                "vat": recipient_vat,
                                "is_company": recipient_is_company,
                                "invoice_number": invoice_number_val,
                                "currency": currency_val,
                                "bank_choice": bank_val,
                                "due_date": due_date_val,
                                "description": description_val,
                                "dataframe": group["dataframe"],
                                "trip_count": group["trip_count"],
                                "total_dkk": group["total_dkk"],
                            })

                    if preview_rows:
                        st.markdown("### Bulk Overview")
                        preview_excel(pd.DataFrame(preview_rows))

                    if st.button("Generate Selected Bulk Invoices", key="generate_bulk_invoices"):
                        validation_errors = validate_bulk_rows(bulk_rows)

                        if validation_errors:
                            for error in validation_errors:
                                st.error(error)
                        else:
                            results = []

                            for row in bulk_rows:
                                if not row["include"]:
                                    continue

                                receiver_dict = {
                                    "name": row["recipient_name"],
                                    "address": row["address"],
                                    "email": row["email"],
                                    "contact": row["contact"],
                                    "vat": row["vat"],
                                    "is_company": row["is_company"],
                                }

                                package = generate_single_invoice_package(
                                    receiver_dict=receiver_dict,
                                    invoice_number=row["invoice_number"],
                                    currency=row["currency"],
                                    bank_choice=row["bank_choice"],
                                    invoice_purpose=row["description"],
                                    due_date=row["due_date"],
                                    cleaned_df=row["dataframe"],
                                )

                                result = {
                                    "group_name": row["group_name"],
                                    "recipient_name": row["recipient_name"],
                                    "recipient_email": row["email"],
                                    "invoice_number": row["invoice_number"],
                                    "currency": row["currency"],
                                    "bank_choice": row["bank_choice"],
                                    "due_date": row["due_date"],
                                    "trip_count": row["trip_count"],
                                    "total_dkk": row["total_dkk"],
                                    "marked_for_sending": row["send_email"],
                                    "send_selected_now": row["send_email"],
                                    **package,
                                }
                                results.append(result)

                            st.session_state.bulk_results = results
                            st.session_state.bulk_zip_bytes = create_zip_from_bulk_results(results) if results else None
                            st.session_state.bulk_zip_name = "bulk_invoices.zip"
                            st.session_state.bulk_last_run_summary = {
                                "generated_count": len(results),
                                "sent_count": 0,
                                "send_failures": [],
                            }

                            if results:
                                st.success(f"Generated {len(results)} invoice package(s).")
                            else:
                                st.warning("No invoices were generated.")

                    if st.session_state.bulk_results:
                        st.markdown("### Generated Bulk Invoices")

                        summary_rows = []
                        for result in st.session_state.bulk_results:
                            summary_rows.append({
                                "Recipient": result["recipient_name"],
                                "Invoice Number": result["invoice_number"],
                                "Trips": result["trip_count"],
                                "Total DKK": round(result["total_dkk"], 2),
                                "Currency": result["currency"],
                                "Email": result["recipient_email"],
                            })

                        preview_excel(pd.DataFrame(summary_rows))

                        if st.session_state.bulk_zip_bytes is not None:
                            st.download_button(
                                "⬇️ Download ZIP with All Invoices",
                                data=st.session_state.bulk_zip_bytes,
                                file_name=st.session_state.bulk_zip_name,
                                mime="application/zip",
                                key="download_bulk_zip"
                            )

                        for idx, result in enumerate(st.session_state.bulk_results):
                            st.markdown(f"#### {result['recipient_name']} — Invoice {result['invoice_number']}")
                            st.download_button(
                                f"⬇️ Download PDF — {result['invoice_number']}",
                                data=result["pdf_bytes"],
                                file_name=result["pdf_filename"],
                                mime="application/pdf",
                                key=f"bulk_pdf_{idx}"
                            )
                            st.download_button(
                                f"⬇️ Download Specification — {result['invoice_number']}",
                                data=result["spec_bytes"],
                                file_name=result["spec_filename"],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"bulk_spec_{idx}"
                            )

                        st.markdown("### Choose Which Generated Invoices to Send")

                        send_selection_rows = []
                        for idx, result in enumerate(st.session_state.bulk_results):
                            col1, col2 = st.columns([4, 1])
                            with col1:
                                send_now = st.checkbox(
                                    f"{result['recipient_name']} — {result['recipient_email']} — Invoice {result['invoice_number']}",
                                    value=result.get("send_selected_now", result.get("marked_for_sending", False)),
                                    key=f"send_select_generated_{idx}"
                                )
                                result["send_selected_now"] = send_now
                            with col2:
                                st.write(result["currency"])

                            send_selection_rows.append({
                                "Send": result["send_selected_now"],
                                "Recipient": result["recipient_name"],
                                "Email": result["recipient_email"],
                                "Invoice Number": result["invoice_number"],
                                "Currency": result["currency"],
                            })

                        if send_selection_rows:
                            preview_excel(pd.DataFrame(send_selection_rows))

                        if st.button("Send Selected Emails", key="send_all_selected_bulk"):
                            send_failures = []
                            send_success_count = 0
                            selected_count = 0

                            for result in st.session_state.bulk_results:
                                if not result.get("send_selected_now", False):
                                    continue

                                selected_count += 1

                                if not result["recipient_email"]:
                                    send_failures.append(
                                        f"{result['recipient_name']} ({result['invoice_number']}): Missing recipient email."
                                    )
                                    continue

                                try:
                                    email_subject = f"Invoice {result['invoice_number']}"
                                    email_body = build_email_body(
                                        customer_name=result["recipient_name"],
                                        invoice_number=result["invoice_number"],
                                        due_date=result["due_date"],
                                        currency=result["currency"]
                                    )
                                    send_email_gmail(
                                        to_email=result["recipient_email"],
                                        subject=email_subject,
                                        body=email_body,
                                        attachments=[
                                            {
                                                "filename": result["pdf_filename"],
                                                "content": result["pdf_bytes"],
                                                "maintype": "application",
                                                "subtype": "pdf",
                                            },
                                            {
                                                "filename": result["spec_filename"],
                                                "content": result["spec_bytes"],
                                                "maintype": "application",
                                                "subtype": "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            },
                                        ]
                                    )
                                    send_success_count += 1
                                except Exception as e:
                                    send_failures.append(
                                        f"{result['recipient_name']} ({result['invoice_number']}): {e}"
                                    )

                            st.session_state.bulk_last_run_summary = {
                                "generated_count": len(st.session_state.bulk_results),
                                "sent_count": send_success_count,
                                "send_failures": send_failures,
                            }

                            if selected_count == 0:
                                st.warning("No invoices were selected for sending.")
                            else:
                                st.success(f"Sent {send_success_count} email(s).")
                                for failure in send_failures:
                                    st.error(failure)

            except Exception as e:
                st.error(f"Could not process bulk Excel file: {e}")

with tab2:
    st.subheader("Manage Customers")

    customers = get_customers()
    if not customers:
        st.info("No customers found. Please add one below.")
    else:
        selected = st.selectbox("Select Customer to Edit/Delete", customers, format_func=lambda c: c.name)

        if selected:
            with st.expander("Edit Customer"):
                name = st.text_input("Name", selected.name)
                email = st.text_input("Email", selected.email)
                address = st.text_input("Address", selected.address)
                contact = st.text_input("Contact", selected.contact)
                vat = st.text_input("VAT Number", selected.vat)
                is_company = st.checkbox("Is Company", selected.is_company)

                if st.button("Update Customer"):
                    update_customer(selected.id, {
                        "name": name,
                        "email": email,
                        "address": address,
                        "contact": contact,
                        "vat": vat,
                        "is_company": is_company
                    })
                    st.success("Customer updated successfully.")

            if st.button("Delete Customer"):
                delete_customer(selected.id)
                st.success("Customer deleted successfully.")

        st.markdown("---")

    with st.expander("Add New Customer"):
        new_name = st.text_input("New Name")
        new_email = st.text_input("New Email")
        new_address = st.text_input("New Address")
        new_contact = st.text_input("New Contact")
        new_vat = st.text_input("New VAT Number")
        new_is_company = st.checkbox("New Is Company", value=True)

        if st.button("Add Customer"):
            if new_name:
                add_customer(
                    name=new_name,
                    email=new_email,
                    address=new_address,
                    contact=new_contact,
                    vat=new_vat,
                    is_company=new_is_company
                )
                st.success("Customer added successfully.")
            else:
                st.warning("Name is required.")
